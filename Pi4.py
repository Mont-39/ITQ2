import sys
import os
import tkinter as tk
import customtkinter as ctk
from PIL import Image, ImageTk
from openpyxl import load_workbook, Workbook
from tkinter import ttk, filedialog
from tkinter import messagebox
from datetime import datetime
import paho.mqtt.client as mqtt

current_dir = os.path.dirname(os.path.abspath(__file__))
bg_dir = os.path.join(current_dir, 'images', 'BGredblue.png')
logo_dir = os.path.join(current_dir, 'images', 'telu.png')
icon_dir = os.path.join(current_dir, 'images', 'taekwondo.xbm')
connect_dir = os.path.join(current_dir, 'images', 'connect.png')
disconnect_dir = os.path.join(current_dir, 'images', 'disconnect.png')

class MQTTHandler():
    def __init__(self, host, port, topics, callback=None):
        self.MQTT_HOST = host
        self.MQTT_PORT = port
        self.MQTT_KEEPALIVE_INTERVAL = 5
        self.message1="1"
        self.message2="2"
        self.message3="3"
        self.message11="1"
        self.message22="2"
        self.message33="3"
        self.topics = topics
        self.received_messages = {}
        self.callback = callback

        self.mqttc = mqtt.Client()
        self.mqttc.on_message = self.on_message
        self.mqttc.on_connect = self.on_connect
        
        self.mqttc.connect(self.MQTT_HOST, self.MQTT_PORT, self.MQTT_KEEPALIVE_INTERVAL)
        self.mqttc.loop_start()

    def on_connect(self, mosq, obj, flags, rc):
        for topic in self.topics:
            self.mqttc.subscribe(topic, 0)
            print(f"Subscribed to MQTT Topic: {topic}")

    def on_message(self, mosq, obj, msg):
        topic = msg.topic
        payload = msg.payload.decode()
        print(f"Received message on topic '{topic}': {payload}")
        self.received_messages[topic] = payload
        self.hand_msg(topic, payload)
        if self.callback:
            self.callback()
        
    def hand_msg(self, topic, payload):
        if topic == "taekwondo/alpha":
            self.message1 = payload
            pass
        elif topic == "taekwondo/beta":
            self.message2 = payload
            pass
        elif topic == "taekwondo/gamma":
            self.message3 = payload
            pass
        elif topic == "taekwondo/alphaa":
            self.message11 = payload
            pass
        elif topic == "taekwondo/betaa":
            self.message22 = payload
            pass
        elif topic == "taekwondo/gammaa":
            self.message33 = payload
            pass

class MyApp():
    def __init__(self):
        self.root = ctk.CTk()
        self.root.geometry("1920x1080")
        #self.root.geometry("1280x720")
        #self.root.after(0, lambda:self.root.state('zoomed'))
        #self.root.resizable(width=False, height=False)
        self.root.title("Operator | Digital Scoring Taekwondo")
        #self.root.iconbitmap(icon_dir)
        
        mqtt_topics = ["taekwondo/alpha", "taekwondo/beta", "taekwondo/gamma", "taekwondo/alphaa", "taekwondo/betaa", "taekwondo/gammaa"]
        self.mqtt_handler = MQTTHandler(host="10.42.0.1", port=1883, topics=mqtt_topics, callback=self.check_msg)

        self.time_left = 120
        self.time_left_display = tk.StringVar(value="2:00")
        self.cd_state = False
        self.label_end = tk.Label(self.root)
        self.label_end = None
        self.label_round = tk.Label(self.root)
        self.label_round = None
        self.label_win1 = tk.Label(self.root)
        self.label_win1 = None
        self.label_win2 = tk.Label(self.root)
        self.label_win2 = None
        self.blinktext()
        self.update_time()

        self.selected_row = None
        self.selected_row_index = None

        bg_img = Image.open(bg_dir)
        self.bgimg = ImageTk.PhotoImage(bg_img)

        self.string1 = tk.StringVar()
        self.string2 = tk.StringVar()
        self.string3 = tk.StringVar()
        self.string11 = tk.StringVar()
        self.string22 = tk.StringVar()
        self.string33 = tk.StringVar()
        self.file_loaded = False
        self.data = None
        self.scoreboard = None
        self.matchlist_win = None
        self.part_state = False
        self.indicator_state = False
        self.end_state = 0
        self.tot_p1 = 0
        self.tot_p2 = 0
        self.currentround = 1
        self.currentwin_roundr = 0
        self.currentwin_roundb = 0
        self.win_roundr = tk.StringVar()
        self.win_roundb = tk.StringVar()
        self.currentpoin1 = 0
        self.currentpoin2 = 0
        self.currentgam1 = 0
        self.currentgam2 = 0
        self.currentkyo1 = 0
        self.currentkyo2 = 0
        self.kyo1 = tk.StringVar()
        self.kyo2 = tk.StringVar()
        self.gam1 = tk.StringVar()
        self.gam2 = tk.StringVar()
        self.round = tk.StringVar()
        self.winner = tk.StringVar()
        self.kategori = tk.StringVar()
        self.nama1 = tk.StringVar()
        self.poin1 = tk.StringVar()
        self.dojang1 = tk.StringVar()
        self.nama2 = tk.StringVar()
        self.poin2 = tk.StringVar()
        self.dojang2 = tk.StringVar()
        
        self.check_points()
        self.win_count()
        self.update_info()

        matchlist_menu = tk.Menu(self.root, background="#39bdd1")
        self.root.config(menu=matchlist_menu)
        file_menu = tk.Menu(matchlist_menu, tearoff=False, background="#39bdd1")
        matchlist_menu.add_cascade(label="Menu", menu=file_menu)
        file_menu.add_command(label="Open Log Participant", command=self.open_log_file)
        file_menu.add_command(label="Open Input Menu", command=self.open_input_menu)
        file_menu.add_separator()
        file_menu.add_command(label="Indicator Menu", command=self.open_indicator)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.exit_anapp)

        self.bg_label = tk.Label(self.root, image=self.bgimg)
        self.bg_label.place(x=0, y=0, relheight=1, relwidth=1)

        label_redround = tk.Label(self.root, textvariable=self.win_roundr, font=('verdana', 30), borderwidth=0, relief="solid", bg='red', fg='white', compound='center')
        label_redround.place(relx=0.04, rely=0.04, anchor='center')

        label_redwin = tk.Label(self.root, text='Round\nWin', font=('verdana', 20), borderwidth=0, relief="solid", bg='red', fg='white', compound='center')
        label_redwin.place(relx=0.04, rely=0.1, anchor='center')

        label_blueround = tk.Label(self.root, textvariable=self.win_roundb, font=('verdana', 30), borderwidth=0, relief="solid", bg='#0022ff', fg='white', compound='center')
        label_blueround.place(relx=0.96, rely=0.04, anchor='center')

        label_bluewin = tk.Label(self.root, text='Round\nWin', font=('verdana', 20), borderwidth=0, relief="solid", bg='#0022ff', fg='white', compound='center')
        label_bluewin.place(relx=0.96, rely=0.1, anchor='center') 

        label_redname = tk.Label(self.root, textvariable=self.nama1, font=('verdana', 30), borderwidth=0, relief="solid", bg='red', fg='white', compound='center')
        label_redname.place(relx=0.21, rely=0.08, anchor='center')

        label_bluename = tk.Label(self.root, textvariable=self.nama2, font=('verdana', 30), borderwidth=0, relief="solid", bg='#0022ff', fg='white', compound='center')
        label_bluename.place(relx=0.75, rely=0.08, anchor='center')

        label_redpoin = tk.Label(self.root, textvariable=self.poin1, font=('verdana', 180), bg='red', fg='white')
        label_redpoin.place(relx=0.15, rely=0.25, anchor='nw')
        
        label_addgam1 = tk.Label(self.root, text="Gam-Jeom\n\n", font=('verdana', 20), bg='red', fg='white')
        label_addgam1.grid(row=1, column=1, padx=2, sticky='s')
        
        label_redgam = tk.Label(self.root, textvariable=self.gam1, font=('verdana', 27), bg='red', fg='white')
        label_redgam.grid(row=1, column=1, padx=2, sticky='s')

        label_addkyo1 = tk.Label(self.root, text="Kyong-Go\n\n", font=('verdana', 20), bg='red', fg='white')
        label_addkyo1.grid(row=1, column=0, padx=2, sticky='s')

        label_redkyo = tk.Label(self.root, textvariable=self.kyo1, font=('verdana', 27), bg='red', fg='white')
        label_redkyo.grid(row=1, column=0, padx=2, sticky='s')

        label_bluepoin = tk.Label(self.root, textvariable=self.poin2, font=('verdana', 180), bg='#0022ff', fg='white')
        label_bluepoin.place(relx=0.75, rely=0.25, anchor='nw')

        label_addgam2 = tk.Label(self.root, text="Gam-Jeom\n\n", font=('verdana', 20), bg='#0022ff', fg='white')
        label_addgam2.grid(row=1, column=8, padx=2, sticky='s')

        label_bluegam = tk.Label(self.root, textvariable=self.gam2, font=('verdana', 27), bg='#0022ff', fg='white')
        label_bluegam.grid(row=1, column=8, padx=2, sticky='s')

        label_addkyo2 = tk.Label(self.root, text="Kyong-Go\n\n", font=('verdana', 20), bg='#0022ff', fg='white')
        label_addkyo2.grid(row=1, column=9, padx=2, sticky='s')

        label_bluekyo = tk.Label(self.root, textvariable=self.kyo2, font=('verdana', 27), bg='#0022ff', fg='white')
        label_bluekyo.grid(row=1, column=9, padx=2, sticky='s')
        
        label_round = tk.Label(self.root, textvariable=self.round, font=("consolas", 39, "bold"))
        label_round.place(relx=0.457, rely=0.2, anchor='nw')
        
        label_time_left = tk.Label(self.root, textvariable=self.time_left_display, font=("consolas", 39, "bold"))
        label_time_left.place(relx=0.464, rely=0.48, anchor='nw')
        
        for i in range(1, 10):
            self.root.grid_columnconfigure(i, weight=1)
        self.root.grid_rowconfigure(0, weight=10)
        empty_label = tk.Label(self.root, bg='red')
        empty_label.grid(row=0)

        self.board_button = tk.Button(self.root, text="Score Board", font=('verdana',10), fg="white", bg="#252525", command=self.displayp, width=20, height=2)
        self.board_button.grid(row=3, column=6, pady=4, padx=2)
        self.file_button = tk.Button(self.root, text="Match Schedule", font=('verdana',10), fg="white", bg="#252525", command=self.matchlist, width=20, height=2)
        self.file_button.grid(row=3, column=5, pady=4, padx=2)
        self.button_start = tk.Button(self.root, text="Start/Pause", font=('verdana',10), fg="white", bg="#252525", command=self.start_pause, width=20, height=2)
        self.button_start.grid(row=3, column=4, pady=4, padx=9)
        self.button_reset = tk.Button(self.root, text="Reset", font=('verdana',10), fg="white", bg="#252525", command=self.reset, width=20, height=2)
        self.button_reset.grid(row=3, column=3, pady=4, padx=7)
        
        self.button_addg1 = tk.Button(self.root, text="+ Gam-jeom", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_gam(1, 1), width=20, height=2)
        self.button_addg1.grid(row=2, column=1, pady=2, padx=3)
        self.button_dedg1 = tk.Button(self.root, text="- Gam-jeom", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_gam(1, -1), width=20, height=2)
        self.button_dedg1.grid(row=3, column=1, pady=2, padx=3)
        self.button_addk1 = tk.Button(self.root, text="+ Kyong-go", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_kyo(1, 1), width=20, height=2)
        self.button_addk1.grid(row=2, pady=2, padx=3)
        self.button_dedk1 = tk.Button(self.root, text="- Kyong-go", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_kyo(1, -1), width=20, height=2)
        self.button_dedk1.grid(row=3, pady=2, padx=3)
        
        self.button_addp1 = tk.Button(self.root, text="+ Poin", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_points(1, 1), width=20, height=2)
        self.button_addp1.grid(row=2, column=2, pady=4, padx=5)
        self.button_dedp1 = tk.Button(self.root, text="- Poin", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_points(1, -1), width=20, height=2)
        self.button_dedp1.grid(row=3, column=2, pady=4, padx=5)
        
        self.button_winred = tk.Button(self.root, text="Win (Red)", font=('verdana',10), fg="white", bg="#252525", command=lambda: self.add_labelwin(1) , width=20, height=2)
        self.button_winred.grid(row=2, column=3, pady=4, padx=7)
        self.button_mdone = tk.Button(self.root, text="Match Done", font=('verdana',10), fg="white", bg="#252525", command=self.delete_selected_row, width=20, height=2)
        self.button_mdone.grid(row=2, column=4, pady=4, padx=9)
        self.button_nround = tk.Button(self.root, text="Next Round", font=('verdana',10), fg="white", bg="#252525", command=lambda: self.change_rounds(1), width=20, height=2)
        self.button_nround.grid(row=2, column=5, pady=4, padx=5)
        self.button_winred = tk.Button(self.root, text="Win (Blue)", font=('verdana',10), fg="white", bg="#252525", command=lambda: self.add_labelwin(2) , width=20, height=2)
        self.button_winred.grid(row=2, column=6, pady=4, padx=5)
        
        self.button_addp2 = tk.Button(self.root, text="+ Poin", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_points(2, 1), width=20, height=2)
        self.button_addp2.grid(row=2, column=7, pady=4, padx=5)
        self.button_dedp2 = tk.Button(self.root, text="- Poin", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_points(2, -1), width=20, height=2)
        self.button_dedp2.grid(row=3, column=7, pady=4, padx=5)
        
        self.button_addg2 = tk.Button(self.root, text="+ Gam-jeom", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_gam(2, 1), width=20, height=2)
        self.button_addg2.grid(row=2, column=8, pady=2, padx=3)
        self.button_dedg2 = tk.Button(self.root, text="- Gam-jeom", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_gam(2, -1), width=20, height=2)
        self.button_dedg2.grid(row=3, column=8, pady=2, padx=3)
        self.button_addk2 = tk.Button(self.root, text="+ Kyong-go", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_kyo(2, 1), width=20, height=2)
        self.button_addk2.grid(row=2, column=9, pady=2, padx=3)
        self.button_dedk2 = tk.Button(self.root, text="- Kyong-go", font=('verdana',9), fg="white", bg="#252525", command=lambda: self.change_kyo(2, -1), width=20, height=2)
        self.button_dedk2.grid(row=3, column=9, pady=2, padx=3)
        
        self.check_points()

    def displayp(self):
        if self.scoreboard is None or not self.scoreboard.winfo_exists():
            self.scoreboard = ctk.CTkToplevel(self.root)
            self.scoreboard.geometry("1280x720")
            self.scoreboard.attributes('-fullscreen', True)
            self.scoreboard.title("Digital Scoring Taekwondo")
            #self.scoreboard.after(250, lambda: self.scoreboard.iconbitmap(icon_dir))

            image = Image.open(logo_dir)
            resizeimage = image.resize((70, 100))
            logo = ImageTk.PhotoImage(resizeimage)
            blank_img = tk.PhotoImage(width=160, height=100)
            self.scoreboard.after(100, lambda: self.scoreboard.focus_force())

            self.update_info()

            self.scoreboard.grid_columnconfigure(0, weight=1, uniform='column')
            self.scoreboard.grid_columnconfigure(1, weight=1)
            self.scoreboard.grid_columnconfigure(2, weight=1, uniform='column')

            self.scoreboard.grid_rowconfigure(0, weight=1, minsize=10)
            self.scoreboard.grid_rowconfigure(1, weight=1, minsize=100)
            self.scoreboard.grid_rowconfigure(2, weight=1, minsize=40)
            self.scoreboard.grid_rowconfigure(3, weight=1, minsize=20)

            label_redname = tk.Label(self.scoreboard, textvariable=self.nama1, font=('verdana', 30), borderwidth=2, relief="solid", bg='red', fg='white', image=blank_img, compound='center')
            label_redname.grid(row=0, column=0, sticky='nsew')

            label_redpoin = tk.Label(self.scoreboard, textvariable=self.poin1, font=('verdana', 100), borderwidth=2, relief="solid", bg='red', fg='white')
            label_redpoin.grid(row=1, column=0, rowspan=2, sticky='nsew')

            label_reddojang = tk.Label(self.scoreboard, textvariable=self.dojang1, font=('verdana', 20), borderwidth=2, relief="solid", bg='red', fg='white', highlightcolor='white')
            label_reddojang.grid(row=3, column=0, sticky='nsew')

            label_logo = tk.Label(self.scoreboard, image=logo, borderwidth=2, relief="solid", bg='#202020')
            label_logo.image = logo
            label_logo.grid(row=0, column=1, sticky='nsew')

            label_time = tk.Label(self.scoreboard, textvariable=self.time_left_display, font=('verdana', 70), borderwidth=2, relief="solid", bg='#6c6c6c', fg='white')
            label_time.grid(row=1, column=1, sticky='nsew')

            label_round = tk.Label(self.scoreboard, textvariable=self.round, font=('verdana', 30), borderwidth=2, relief="solid", bg='#202020', fg='white')
            label_round.grid(row=2, column=1, sticky='nsew')

            label_weight = tk.Label(self.scoreboard, textvariable=self.kategori, font=('verdana', 20), borderwidth=2, relief="solid", bg='#202020', fg='white')
            label_weight.grid(row=3, column=1, sticky='nsew')

            label_bluename = tk.Label(self.scoreboard, textvariable=self.nama2, font=('verdana', 30), borderwidth=2, relief="solid", bg='blue', fg='white', image=blank_img, compound='center')
            label_bluename.grid(row=0, column=2, sticky='nsew')

            label_bluepoin = tk.Label(self.scoreboard, textvariable=self.poin2, font=('verdana', 100), borderwidth=2, relief="solid", bg='blue', fg='white')
            label_bluepoin.grid(row=1, column=2, rowspan=2, sticky='nsew')

            label_bluedojang = tk.Label(self.scoreboard, textvariable=self.dojang2, font=('verdana', 20), borderwidth=2, relief="solid", bg='blue', fg='white')
            label_bluedojang.grid(row=3, column=2, sticky='nsew')
        else:
            self.close_scoreboard()

    def matchlist(self):
        if self.matchlist_win is None or not self.matchlist_win.winfo_exists():
            self.matchlist_win = ctk.CTkToplevel(self.root, fg_color="#39bbd1")
            self.matchlist_win.geometry("1280x300")
            self.matchlist_win.title("Match List | Digital Scoring Taekwondo")
            self.matchlist_win.after(250, lambda: self.matchlist_win.iconbitmap(icon_dir))
            
            self.l_frame = tk.Frame(self.matchlist_win, bg="#39bbd1")
            self.l_frame.pack(side='left', fill='y')
            self.b_frame = tk.Frame(self.matchlist_win)
            self.b_frame.pack(anchor="n", side="bottom", fill='none', expand=True)
            custom_font=("Helvetica", 12, "italic bold")

            self.label_entry1 = tk.Label(self.l_frame, text="Chose Match\nby click the row", font=custom_font, bg="#39bbd1")
            self.label_entry1.pack(pady=2)

            self.load_button = tk.Button(self.l_frame, bg="#2b8d9e", text="Refresh Table", command=self.load_file)
            self.load_button.pack(pady=5)

            self.delete_button = tk.Button(self.l_frame, bg="#2b8d9e", text="Delete Selected Row", command=self.delete_selected_row)
            self.delete_button.pack(pady=10)

            self.matchlist_win.after(100, lambda:
                                     self.matchlist_win.focus_force())

            self.r_frame = tk.Frame(self.matchlist_win, bg='teal')
            self.r_frame.pack(side='right')

            style = ttk.Style()
            style.theme_use("classic")
            style.configure('Treeview.Heading', background='#39bbd1', foreground='black')
            self.tree = ttk.Treeview(self.r_frame, selectmode="extended", show="headings")
            self.tree.pack(fill="both", expand=True)
            self.tree.tag_configure('oddrow', background="#2b8d9e")
            self.tree.tag_configure('evenrow', background="#39bbd1")
            self.tree.bind("<<TreeviewSelect>>", self.on_row_select)
            self.tree.bind("<<TreeviewSelect>>", self.participant_state() ,add="+")

            if self.file_loaded and self.data:
                self.table()
        else:
            self.close_matchlist()

    def participant_state(self):
        self.part_state = True

    def clear_treeview(self):
        self.tree.delete(*self.tree.get_children())

    def on_row_select(self, event):
        selected_item = self.tree.selection()[0]
        self.selected_row = self.tree.item(selected_item, "values")
        print(f"selected row: {self.selected_row}")
        print(f"selected item: {selected_item}")
        self.selected_row_index = self.tree.index(selected_item) + 2
        self.kategori.set("Under\n"+self.selected_row[0])
        self.nama1.set(self.selected_row[1])
        self.dojang1.set(self.selected_row[4])
        self.nama2.set(self.selected_row[5])
        self.dojang2.set(self.selected_row[8])
        self.done_label = tk.Label(self.matchlist_win, text="Done")
        self.done_label.pack()

    def save_deleted_row(self, row, format="%Y-%m-%d %H:%M:%S"):
        round_red = self.currentwin_roundr
        round_blue = self.currentwin_roundb
        timestamp = datetime.now().strftime(format)
        log_filename = "participant_log.xlsx"
        try:
            log_wb = load_workbook(log_filename)
            log_ws = log_wb.active
        except FileNotFoundError:
            log_wb = Workbook()
            log_ws = log_wb.active
            log_ws.append(["Timestamp", "Rounds Win(R)", "Rounds Win(B)"] + [cell for cell in self.tree["column"]])
        
        log_ws.append([timestamp, round_red, round_blue] + list(row))
        log_wb.save(log_filename)

    def delete_selected_row(self):
        if self.selected_row_index is not None:
            self.save_deleted_row(self.selected_row)
            self.ws.delete_rows(self.selected_row_index)
            self.wb.save(self.filename)
            self.reset()
            self.clear_treeview()
            self.table()

    def open_log_file(self):
        log_filename = "participant_log.xlsx"
        try:
            log_wb = load_workbook(log_filename)
            log_ws = log_wb.active

            log_window = tk.Toplevel(self.root)
            log_window.title("Concluded Match")
            log_window.geometry("1000x500")
            log_window.configure(background="#39bdd1")
            
            log_canvas = tk.Canvas(log_window, bg="#39bdd1")
            log_canvas.pack(fill="both", expand=True)
            
            log_tree = ttk.Treeview(log_canvas)
            log_tree.pack(side="left", fill="both", expand=True)

            log_scrollbar_y = tk.Scrollbar(log_canvas, orient="vertical", command=log_tree.yview)
            log_scrollbar_y.pack(side="right", fill="y")
            log_tree.configure(yscrollcommand=log_scrollbar_y.set)

            log_scrollbar_x = tk.Scrollbar(log_window, orient="horizontal", command=log_tree.xview)
            log_scrollbar_x.pack(side="bottom", fill="x")
            log_tree.configure(xscrollcommand=log_scrollbar_x.set)

            log_tree["column"] = [cell.value for cell in log_ws[1]]
            log_tree["show"] = "headings"
            for col in log_tree["column"]:
                log_tree.column(col, stretch=True)
            for col in log_tree["column"]:
                log_tree.heading(col, text=col, anchor="w")
            for row in log_ws.iter_rows(min_row=2, values_only=True):
                log_tree.insert("", "end", values=row)
        except FileNotFoundError:
            messagebox.showwarning("File not Found", "Is there any concluded match?")
            print("Log file not found.")
    
    def open_input_menu(self):
        input_window = tk.Toplevel(self.root)
        input_window.title("Input Data Menu")

        tk.Label(input_window, text="Category").grid(row=0, column=0, columnspan=1)
        self.entry_category = tk.Entry(input_window)
        self.entry_category.grid(row=0, column=1, columnspan=2)

        frame1 = tk.Frame(input_window)
        frame1.grid(row=1, column=0, columnspan=2, pady=10)

        tk.Label(frame1, text="Name 1").grid(row=0, column=0)
        self.entry_name1 = tk.Entry(frame1)
        self.entry_name1.grid(row=0, column=1)

        tk.Label(frame1, text="Weight 1").grid(row=1, column=0)
        self.entry_weight1 = tk.Entry(frame1)
        self.entry_weight1.grid(row=1, column=1)

        tk.Label(frame1, text="Gender 1").grid(row=2, column=0)
        self.entry_gender1 = tk.Entry(frame1)
        self.entry_gender1.grid(row=2, column=1)

        tk.Label(frame1, text="Dojang 1").grid(row=3, column=0)
        self.entry_dojang1 = tk.Entry(frame1)
        self.entry_dojang1.grid(row=3, column=1)

        frame2 = tk.Frame(input_window)
        frame2.grid(row=1, column=2, columnspan=2, pady=10)

        tk.Label(frame2, text="Name 2").grid(row=0, column=0)
        self.entry_name2 = tk.Entry(frame2)
        self.entry_name2.grid(row=0, column=1)

        tk.Label(frame2, text="Weight 2").grid(row=1, column=0)
        self.entry_weight2 = tk.Entry(frame2)
        self.entry_weight2.grid(row=1, column=1)

        tk.Label(frame2, text="Gender 2").grid(row=2, column=0)
        self.entry_gender2 = tk.Entry(frame2)
        self.entry_gender2.grid(row=2, column=1)

        tk.Label(frame2, text="Dojang 2").grid(row=3, column=0)
        self.entry_dojang2 = tk.Entry(frame2)
        self.entry_dojang2.grid(row=3, column=1)

        save_button = tk.Button(input_window, text="Save to Excel", command=self.save_to_excel)
        save_button.grid(row=2, columnspan=4, pady=10)

    def save_to_excel(self):
        category = self.entry_category.get()
        name1 = self.entry_name1.get()
        weight1 = self.entry_weight1.get()
        gender1 = self.entry_gender1.get()
        dojang1 = self.entry_dojang1.get()

        name2 = self.entry_name2.get()
        weight2 = self.entry_weight2.get()
        gender2 = self.entry_gender2.get()
        dojang2 = self.entry_dojang2.get()

        if not all([category, name1, weight1, gender1, dojang1, name2, weight2, gender2, dojang2]):
            messagebox.showwarning("Input Error", "Semua data harus terisi")
            return

        if os.path.exists("participant.xlsx"):
            wb = load_workbook("participant.xlsx")
        else:
            wb = Workbook()

        ws = wb.active
        ws.title = "Participant"
        max_row = 1

        if ws.max_row == 1 and ws['A1'].value is None:
            ws.cell(row=max_row, column=1, value="Weight Category")
            ws.cell(row=max_row, column=2, value="Name(red)")
            ws.cell(row=max_row, column=3, value="weight(red)")
            ws.cell(row=max_row, column=4, value="gender(red)")
            ws.cell(row=max_row, column=5, value="dojang(red)")
            ws.cell(row=max_row, column=6, value="name(blue)")
            ws.cell(row=max_row, column=7, value="weight(blue)")
            ws.cell(row=max_row, column=8, value="gender(blue)")
            ws.cell(row=max_row, column=9, value="dojang(blue)")

        next_row = ws.max_row + 1

        ws.cell(row=next_row, column=1, value=category)
        ws.cell(row=next_row, column=2, value=name1)
        ws.cell(row=next_row, column=3, value=weight1)
        ws.cell(row=next_row, column=4, value=gender1)
        ws.cell(row=next_row, column=5, value=dojang1)
        ws.cell(row=next_row, column=6, value=name2)
        ws.cell(row=next_row, column=7, value=weight2)
        ws.cell(row=next_row, column=8, value=gender2)
        ws.cell(row=next_row, column=9, value=dojang2)

        wb.save("participant.xlsx")
        self.message_input()

    def message_input(self):
        messagebox.showinfo("Success", "Data saved to Excel file successfully")
        self.table()

    def chose_row(self):
        self.part_state = True
        row_number = int(self.row_entry1.get())

        self.kategori.set(self.data[row_number][0])
        self.nama1.set(self.data[row_number][1])
        self.dojang1.set(self.data[row_number][4])
        self.nama2.set(self.data[row_number][5])
        self.dojang2.set(self.data[row_number][8])

        self.done_label = tk.label(self.l_frame, text="Done")
        self.done_label.pack()

    def load_file(self):
        self.filename = ("participant.xlsx")
        if self.filename:
            try:
                self.wb = load_workbook(self.filename)
                self.ws = self.wb.active
            except Exception as e:
                print(f"Error: {e}")
        self.data = [[cell.value for cell in row] for row in self.ws.iter_rows()]
        self.file_loaded = True
        self.table()

    def table(self):
        self.clear_treeview()
        self.tree["column"] = [cell.value for cell in self.ws[1]]
        self.tree["show"] = "headings"
        for col in self.tree["column"]:
            col_width = max(len(str(value)) for value in self.ws[1])
            self.tree.column(col, width=col_width * 10, anchor="center")
            self.tree.heading(col, text=col)

        for i, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=1):
            tag = "oddrow" if i % 2 else "evenrow"
            self.tree.insert("", "end", values=row, tags=(tag,))
            self.tree.pack()
    
    def start_pause(self):
        if self.part_state == True:
            self.end_state = 0
            self.cd_state = not self.cd_state 
        else:
            self.pop_part()

    def reset(self):
        self.part_state = False
        self.currentpoin1 = 0
        self.currentpoin2 = 0
        self.currentkyo1 = 0
        self.currentkyo2 = 0
        self.currentgam1 = 0
        self.currentgam2 = 0
        self.currentround = 1
        self.currentwin_roundr = 0
        self.currentwin_roundb = 0
        self.kategori.set('')
        self.nama1.set('')
        self.dojang1.set('')
        self.nama2.set('')
        self.dojang2.set('')
        self.time_left = 120
        self.time_left_display.set("2:00")
        self.cd_state = False
        self.remove_labelwin()
        self.remove_labelend()
        self.remove_labelround()
        self.update_info()

    def reset_round(self):
        self.currentpoin1 = 0
        self.currentpoin2 = 0
        self.currentkyo1 = 0
        self.currentkyo2 = 0
        self.currentgam1 = 0
        self.currentgam2 = 0
        self.time_left = 120
        self.time_left_display.set("2:00")
        self.cd_state = False
        self.remove_labelround()
        self.update_info()        

    def update_time(self):
        if self.cd_state and self.time_left > 0:
            self.time_left -= 1
            mins, secs = divmod(self.time_left, 60)
            self.time_left_display.set(f"{mins}:{secs:02d}")
            if self.time_left == 0:
                self.round_win()
                self.add_labelend()
                self.cd_state = not self.cd_state
        self.root.after(1000, self.update_time)
    
    def add_labelend(self):
        if self.label_end is None:
            self.label_end = tk.Label(self.root, text="Times Up!", font=("verdana", 25, "italic"), foreground='black')
            self.label_end.place(relx=0.5, rely=0.69, anchor='center')
            
            self.label_end1 = tk.Label(self.scoreboard, text="Times Up!", font=("verdana", 25, "italic"), foreground='black')
            self.label_end1.place(relx=0.5, rely=0.69, anchor='center')
            #self.blinktext()

    def remove_labelend(self):
        if self.label_end is not None:
            self.label_end.destroy()
            self.label_end = None
            self.label_end1.destroy()
            self.label_end1 = None

    def add_labelround(self):
        if self.label_round is None:
            self.label_round = tk.Label(self.root, text="Rounds End!", font=("verdana", 25, "italic"), foreground='black')
            self.label_round.place(relx=0.5, rely=0.69, anchor='center')
            #self.blinktext()
            self.label_round1 = tk.Label(self.scoreboard, text="Rounds End!", font=("verdana", 25, "italic"), foreground='black')
            self.label_round1.place(relx=0.5, rely=0.69, anchor='center')

    def remove_labelround(self):
        if self.label_round is not None:
            self.label_round.destroy()
            self.label_round = None
            self.label_round1.destroy()
            self.label_round1 = None

    def add_labelwin(self, teams):
        if teams == 1:
            winner_name = self.nama1.get()
            self.winner.set(f"Winner\n{winner_name}")
            if self.label_win1 is None:
                self.label_win1 = tk.Label(self.root, textvariable=self.winner, font=("verdana", 25, "italic"), foreground='black')
                self.label_win1.place(relx=0.5, rely=0.8, anchor='center')
                self.label_win2 = tk.Label(self.scoreboard, textvariable=self.winner, font=("verdana", 35, "italic"), foreground='black')
                self.label_win2.place(relx=0.5, rely=0.55, anchor='center')
                #self.blinktext()
        elif teams == 2:
            winner_name = self.nama2.get()
            self.winner.set(f"Winner\n{winner_name}")
            if self.label_win1 is None:
                self.label_win1 = tk.Label(self.root, textvariable=self.winner, font=("verdana", 25, "italic"), foreground='black')
                self.label_win1.place(relx=0.5, rely=0.8, anchor='center')
                self.label_win2 = tk.Label(self.scoreboard, textvariable=self.winner, font=("verdana", 35, "italic"), foreground='black')
                self.label_win2.place(relx=0.5, rely=0.55, anchor='center')
                #self.blinktext()

    def remove_labelwin(self):
        if self.label_win1 is not None:
            self.label_win1.destroy()
            self.label_win1 = None
            self.label_win2.destroy()
            self.label_win2 = None

    def blinktext(self):
        blink_fast = 1000
        blink_slow = 2000

        if self.label_round is not None:
            self.colour_lround = self.label_round.cget("foreground")
            if self.colour_lround == "white":
                self.label_round.config(foreground="black")
                self.root.after(blink_slow, self.blinktext)
            else:
                self.label_round.config(foreground="white")
                self.root.after(blink_fast, self.blinktext)

        if self.label_end is not None:
            self.colour_lend = self.label_end.cget("foreground")
            if self.colour_lend == "white":
                self.label_end.config(foreground="black")
                self.root.after(blink_slow, self.blinktext)
            else:
                self.label_end.config(foreground="white")
                self.root.after(blink_fast, self.blinktext)
        if self.label_win1 is not None:
            self.colour_lwin1 = self.label_win1.cget("foreground")
            if self.colour_lwin1 == "white":
                self.label_win1.config(foreground="black")
                self.label_win2.config(foreground="black")
                self.root.after(blink_slow, self.blinktext)
            else:
                self.label_win1.config(foreground="white")
                self.label_win2.config(foreground="white")
                self.root.after(blink_fast, self.blinktext)

    def change_points_mqtt(self, team, increment):
        if self.cd_state:
            if team == 1:
                self.currentpoin1 += increment
                if self.currentpoin1 < 0:
                    self.currentpoin1 = 0
                self.poin1.set(f"{self.currentpoin1}")
                self.check_points()
            elif team == 2:
                self.currentpoin2 += increment
                if self.currentpoin2 < 0:
                    self.currentpoin2 = 0
                self.poin2.set(f"{self.currentpoin2}")
                self.check_points()
        else:
            win_pop = tk.Tk()
            win_pop.withdraw()
            response = messagebox.showinfo("Tidak dapat mengubah poin", "Countdown belum berjalan.")
            if response == 'ok':
                win_pop.destroy()

    def change_points(self, team, increment):
        if team == 1:
            self.currentpoin1 += increment
            if self.currentpoin1 < 0:
                self.currentpoin1 = 0
            self.poin1.set(f"{self.currentpoin1}")
            self.check_points()
        elif team == 2:
            self.currentpoin2 += increment
            if self.currentpoin2 < 0:
                self.currentpoin2 = 0
            self.poin2.set(f"{self.currentpoin2}")
            self.check_points()

    def change_gam(self, team, increment):
        if team == 1:
            self.currentgam1 += increment
            if self.currentgam1 < 0:
                self.currentgam1 = 0
            self.gam1.set(f"{self.currentgam1}")
            self.check_points()
        elif team == 2:
            self.currentgam2 += increment
            if self.currentgam2 < 0:
                self.currentgam2 = 0
            self.gam2.set(f"{self.currentgam2}")
            self.check_points()

    def change_kyo(self, team, increment):
        if team == 1:
            self.currentkyo1 += increment
            if self.currentkyo1 < 0:
                self.currentkyo1 = 0
            self.kyo1.set(f"{self.currentkyo1}")
            self.check_points()
        elif team == 2:
            self.currentkyo2 += increment
            if self.currentkyo2 < 0:
                self.currentkyo2 = 0
            self.kyo2.set(f"{self.currentkyo2}")
            self.check_points()
    
    def win_count(self):
        if self.currentwin_roundr == 2:
            self.add_labelwin(1)
        elif self.currentwin_roundb == 2:
            self.add_labelwin(2)
    
    def round_win(self):
        if self.tot_p1 > self.tot_p2:
            self.currentwin_roundr+= 1
            self.win_roundr.set(f"{self.currentwin_roundr}")
            self.update_info()
            self.win_count()
        elif self.tot_p1 < self.tot_p2:
            self.currentwin_roundb += 1
            self.win_roundb.set(f"{self.currentwin_roundb}")
            self.update_info()
            self.win_count()
    
    def check_points(self):
        self.tot_p1 = (self.currentpoin1 + self.currentkyo2 + self.currentgam2)
        self.tot_p2 = (self.currentpoin2 + self.currentkyo1 + self.currentgam1)
        if abs(self.tot_p1 - self.tot_p2) >= 10:
            self.round_win()
            self.start_pause()
            self.add_labelround()

    def change_rounds(self, increment):
        if self.currentround < 3:
            if increment == 1:
                self.currentround += 1
                self.round.set(f"Round\n{self.currentround}")
                self.reset_round()
                self.update_info()

    def update_info(self):
        self.poin1.set(f"{self.currentpoin1}")
        self.poin2.set(f"{self.currentpoin2}")
        self.gam1.set(f"{self.currentgam1}")
        self.gam2.set(f"{self.currentgam2}")
        self.kyo1.set(f"{self.currentkyo1}")
        self.kyo2.set(f"{self.currentkyo2}")
        self.round.set(f"Round\n{self.currentround}")
        self.root.after(100, self.update_info)
        self.win_roundr.set(f"{self.currentwin_roundr}")
        self.win_roundb.set(f"{self.currentwin_roundb}")

    def close_scoreboard(self):
        if self.scoreboard:
            self.scoreboard.destroy()
            self.scoreboard = None
    
    def close_matchlist(self):
        if self.matchlist_win:
            self.matchlist_win.destroy()
            self.matchlist_win = None

    def pop_part(self):
        win_pop = tk.Tk()
        win_pop.withdraw()
        response = messagebox.showinfo("Tidak dapat menjalankan timer", "Pertandingan belum dipilih.")
        if response == 'ok':
            win_pop.destroy()

    def check_msg(self):
        self.toggle_image()
        self.string1 = self.mqtt_handler.message1
        self.string2 = self.mqtt_handler.message2
        self.string3 = self.mqtt_handler.message3
        unique_strings = set([self.string1,
                            self.string2,
                            self.string3])
        print(f"{unique_strings}")
        if len(unique_strings) == 2:
            print("Two message are the same")
            duplicate_strings = [string for string in unique_strings if (self.string1, self.string2, self.string3).count(string) > 1]
            for duplicate in duplicate_strings:
                point = duplicate
                print(f"the same string: {point}")
            if point == "Red1":
                self.change_points_mqtt(1, 1)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"
                print(f"{unique_strings}")
            elif point == "Red2":
                self.change_points_mqtt(1, 2)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"
            elif point == "Red3":
                self.change_points_mqtt(1, 3)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"
            elif point == "Blue1":
                self.change_points_mqtt(2, 1)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"
            elif point == "Blue2":
                self.change_points_mqtt(2, 2)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"
            elif point == "Blue3":
                self.change_points_mqtt(2, 3)
                self.mqtt_handler.message1 = "1"
                self.mqtt_handler.message2 = "2"
                self.mqtt_handler.message3 = "3"           
        else:
            print("Message are unique")

    def open_indicator(self):
        if not self.indicator_state:
            self.indicator_menu = tk.Toplevel(self.root)
            self.indicator_menu.title("Connection Indicator")
            self.indicator_menu.configure(background="white")

            image11 = Image.open(connect_dir)
            image22 = Image.open(disconnect_dir)

            resized11 = image11.resize((300, 300))
            resized22 = image22.resize((300, 300))

            self.image_con = ImageTk.PhotoImage(resized11)
            self.image_dis = ImageTk.PhotoImage(resized22)

            image_width, image_height = 300, 300
            self.image_label1 = tk.Label(self.indicator_menu, image=self.image_dis, width=image_width, height=image_height, anchor="center")
            self.image_label2 = tk.Label(self.indicator_menu, image=self.image_dis, width=image_width, height=image_height, anchor="center")
            self.image_label3 = tk.Label(self.indicator_menu, image=self.image_dis, width=image_width, height=image_height, anchor="center")

            self.image_label1.grid(row=0, column=0, sticky="nsew")
            self.image_label2.grid(row=0, column=1, sticky="nsew")
            self.image_label3.grid(row=0, column=2, sticky="nsew")
            
            label_p1 = tk.Label(self.indicator_menu, text="Alpha State", font=('verdana', 20), bg="white")
            label_p1.grid(row=1, column=0)
            label_p2 = tk.Label(self.indicator_menu, text="Beta State", font=('verdana', 20), bg="white")
            label_p2.grid(row=1, column=1)
            label_p3 = tk.Label(self.indicator_menu, text="Gamma State", font=('verdana', 20), bg="white")
            label_p3.grid(row=1, column=2)
            exit_button = tk.Button(self.indicator_menu, text="Exit", command=self.close_indicator)
            exit_button.grid(row=2, column=1)
            self.indicator_state = True
        else:
            messagebox.showinfo("Info", "Menu Indikator telah dibuka!")

    def close_indicator(self):
        if self.indicator_state:
            self.indicator_menu.destroy()
            self.indicator_state = False
        else:
            messagebox.showinfo("Info", "Menu Indikator Telah tertutup!")

    def toggle_image(self):
        self.string11 = self.mqtt_handler.message11
        self.string22 = self.mqtt_handler.message22
        self.string33 = self.mqtt_handler.message33
        if (self.string11 == "a") & (self.indicator_state == True):
            self.image_label1.config(image=self.image_con)
        if (self.string22 == "a") & (self.indicator_state == True):
            self.image_label2.config(image=self.image_con)
        if (self.string33 == "a") & (self.indicator_state == True):
            self.image_label3.config(image=self.image_con)
            
    def exit_anapp(self):
        sys.exit(0)
    
if __name__ == "__main__":
    app = MyApp()
    app.root.mainloop()    
